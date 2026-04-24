import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from app.services.currency import convert_amount


# ── Palette ─────────────────────────────────────────────────────────────────
PRIMARY = "1A3A5C"        # bleu nuit Valmere
PRIMARY_SOFT = "2B5680"   # bleu intermédiaire (dégradé visuel)
GOLD = "C9A84C"
LIGHT = "F1F5F9"
LIGHTER = "F8FAFC"
WHITE = "FFFFFF"
TEXT_DARK = "1E293B"
TEXT_MUTED = "64748B"
BORDER = "E2E8F0"

# Transaction type colors — zebra tones stay subtle, text colors pop.
TX_TYPE_STYLES = {
    "initial":    {"bg": "EFF6FF", "fg": "1E40AF", "label_fr": "Apport initial", "label_en": "Initial capital",  "label_es": "Capital inicial"},
    "deposit":    {"bg": "EFF6FF", "fg": "1D4ED8", "label_fr": "Dépôt",          "label_en": "Deposit",          "label_es": "Depósito"},
    "withdrawal": {"bg": "FFF7ED", "fg": "C2410C", "label_fr": "Retrait",        "label_en": "Withdrawal",       "label_es": "Retiro"},
    "gain":       {"bg": "ECFDF5", "fg": "047857", "label_fr": "Gain",           "label_en": "Gain",             "label_es": "Ganancia"},
    "loss":       {"bg": "FEF2F2", "fg": "B91C1C", "label_fr": "Perte",          "label_en": "Loss",             "label_es": "Pérdida"},
    "fee":        {"bg": "F5F3FF", "fg": "6D28D9", "label_fr": "Frais",          "label_en": "Fee",              "label_es": "Comisión"},
}


# ── Minimal i18n ────────────────────────────────────────────────────────────
I18N = {
    "fr": {
        "title": "Relevé de compte",
        "generated_at": "Généré le",
        "period": "Période",
        "investor": "Investisseur",
        "code": "Code",
        "summary": "Résumé du portefeuille",
        "initial": "Capital initial",
        "current": "Valeur actuelle",
        "pnl": "Gain / Perte",
        "roi": "Rendement (ROI)",
        "entry_date": "Date d'entrée",
        "status": "Statut",
        "history": "Historique des transactions",
        "date": "Date",
        "type": "Type",
        "amount": "Montant",
        "original": "Montant d'origine",
        "orig_ccy": "Devise d'origine",
        "description": "Description",
        "reference": "Référence",
        "performances": "Performances par période",
        "start": "Début",
        "end": "Fin",
        "roi_pct": "ROI (%)",
        "gross_gain": "Gain brut",
        "drawdown": "Drawdown max",
        "na": "—",
        "legend": "Légende",
        "initial_deposit_desc": "Apport initial à l'ouverture du compte",
    },
    "en": {
        "title": "Account statement",
        "generated_at": "Generated on",
        "period": "Period",
        "investor": "Investor",
        "code": "Code",
        "summary": "Portfolio summary",
        "initial": "Initial capital",
        "current": "Current value",
        "pnl": "Gain / Loss",
        "roi": "Return (ROI)",
        "entry_date": "Entry date",
        "status": "Status",
        "history": "Transaction history",
        "date": "Date",
        "type": "Type",
        "amount": "Amount",
        "original": "Original amount",
        "orig_ccy": "Original currency",
        "description": "Description",
        "reference": "Reference",
        "performances": "Performance by period",
        "start": "Start",
        "end": "End",
        "roi_pct": "ROI (%)",
        "gross_gain": "Gross gain",
        "drawdown": "Max drawdown",
        "na": "—",
        "legend": "Legend",
        "initial_deposit_desc": "Initial deposit at account opening",
    },
    "es": {
        "title": "Estado de cuenta",
        "generated_at": "Generado el",
        "period": "Período",
        "investor": "Inversor",
        "code": "Código",
        "summary": "Resumen de la cartera",
        "initial": "Capital inicial",
        "current": "Valor actual",
        "pnl": "Ganancia / Pérdida",
        "roi": "Rendimiento (ROI)",
        "entry_date": "Fecha de entrada",
        "status": "Estado",
        "history": "Historial de transacciones",
        "date": "Fecha",
        "type": "Tipo",
        "amount": "Monto",
        "original": "Monto original",
        "orig_ccy": "Moneda original",
        "description": "Descripción",
        "reference": "Referencia",
        "performances": "Rendimiento por período",
        "start": "Inicio",
        "end": "Fin",
        "roi_pct": "ROI (%)",
        "gross_gain": "Ganancia bruta",
        "drawdown": "Drawdown máx",
        "na": "—",
        "legend": "Leyenda",
        "initial_deposit_desc": "Aporte inicial a la apertura de la cuenta",
    },
}


def _t(lang: str, key: str) -> str:
    return I18N.get(lang, I18N["fr"]).get(key, key)


def _type_label(t: str, lang: str) -> str:
    meta = TX_TYPE_STYLES.get((t or "").lower())
    if not meta:
        return (t or "").upper()
    return meta.get(f"label_{lang}") or meta.get("label_fr") or t.upper()


def _thin_border(color: str = BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fetch_logo_bytes(url: str) -> bytes | None:
    if not url:
        return None
    try:
        import urllib.request
        with urllib.request.urlopen(url, timeout=5) as r:
            return r.read()
    except Exception:
        return None


def _paint_row(ws, row: int, col_start: int, col_end: int, bg: str):
    """Fill background on a range of cells — keeps the row visually contiguous."""
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=bg)


def generate_statement(
    investor, investment, transactions, performances,
    period_start=None, period_end=None,
    lang: str = "fr",
    display_currency: str = "HTG",
    db=None,
    company=None,
) -> bytes:
    """
    Génère un relevé Excel avec :
      • un entête « bannière » bleu dégradé (logo + nom d'entreprise + titre)
      • un bloc résumé alterné
      • un historique où CHAQUE type de transaction a sa couleur propre
      • le capital initial apparaît comme PREMIÈRE ligne du tableau (type « INITIAL »)
      • une légende des couleurs en fin de document
    """
    wb = Workbook()
    ws = wb.active
    ws.title = _t(lang, "title")[:31]
    ws.sheet_view.showGridLines = False  # plus lisible pour un rapport financier

    # Column widths — laisser de l'espace aux libellés longs et aux montants.
    for col, width in [("A", 14), ("B", 22), ("C", 20), ("D", 20), ("E", 18), ("F", 40)]:
        ws.column_dimensions[col].width = width

    display_ccy = (display_currency or "HTG").upper()
    inv_ccy = (getattr(investment, "currency", None) or "HTG").upper()
    company_name = getattr(company, "company_name", None) or "Valmere & Co"

    # ── Bannière d'en-tête centrée (lignes 1-5) ───────────────────────────
    # Logo centré, puis nom de l'entreprise dessous (également centré),
    # suivi du titre « Relevé de compte » et du sous-titre.
    header_fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
    for r in (1, 2, 3, 4, 5):
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = header_fill

    # Hauteurs des lignes choisies pour laisser respirer le logo.
    ws.row_dimensions[1].height = 6                 # petit espace supérieur
    ws.row_dimensions[2].height = 60                # logo
    ws.row_dimensions[3].height = 26                # nom entreprise
    ws.row_dimensions[4].height = 18                # titre
    ws.row_dimensions[5].height = 18                # sous-titre (date/période)

    # Logo centré horizontalement via un OneCellAnchor avec offset calculé.
    # Largeurs colonnes (A..F) : 14,22,20,20,18,40 ≈ 98,154,140,140,126,280 px ⇒ ~938 px total.
    logo_bytes = _fetch_logo_bytes(company.logo_url) if (company and getattr(company, "logo_url", None)) else None
    if logo_bytes:
        try:
            logo_w, logo_h = 160, 52
            total_w = 98 + 154 + 140 + 140 + 126 + 280  # approx en px
            left_px = (total_w - logo_w) // 2            # centre
            # position : colonne C (indice 2, col start = 98+154 = 252 px)
            col_idx = 2
            col_off_px = left_px - (98 + 154)
            # verticalement : ligne 2 (indice 1). Hauteur 60pt ≈ 80 px ; offset top pour centrer.
            row_idx = 1
            row_off_px = max(0, (80 - logo_h) // 2)

            img = XLImage(io.BytesIO(logo_bytes))
            img.width = logo_w
            img.height = logo_h
            img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=col_idx, colOff=pixels_to_EMU(col_off_px),
                    row=row_idx, rowOff=pixels_to_EMU(row_off_px),
                ),
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(logo_w),
                    cy=pixels_to_EMU(logo_h),
                ),
            )
            ws.add_image(img)
        except Exception:
            pass

    # Nom de l'entreprise — centré, gras, blanc
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = company_name
    c.font = Font(bold=True, size=18, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Titre du document — doré, centré
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = _t(lang, "title").upper()
    c.font = Font(bold=True, size=11, color=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Sous-titre : date de génération + période éventuelle — centré
    ws.merge_cells("A5:F5")
    now_str = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")
    sub = f"{_t(lang, 'generated_at')} · {now_str}"
    if period_start:
        sub += f"   |   {_t(lang, 'period')} · {period_start} → {period_end}"
    c = ws["A5"]
    c.value = sub
    c.font = Font(color="CBD5E1", italic=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Carte « Investisseur » (lignes 7-8) ───────────────────────────────
    # Ligne 6 : espace blanc après la bannière
    ws.row_dimensions[6].height = 8

    ws.merge_cells("A7:F7")
    c = ws["A7"]
    c.value = f"   {_t(lang, 'investor').upper()}"
    c.font = Font(bold=True, size=9, color=TEXT_MUTED, name="Calibri")
    c.fill = PatternFill(fill_type="solid", fgColor=LIGHTER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 16

    ws.merge_cells("A8:F8")
    c = ws["A8"]
    c.value = f"   {investor.full_name}          ·          {_t(lang, 'code')} : {investor.code}"
    c.font = Font(bold=True, size=13, color=PRIMARY)
    c.fill = PatternFill(fill_type="solid", fgColor=LIGHTER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 24
    # bottom border for the card
    for col in range(1, 7):
        ws.cell(row=8, column=col).border = Border(bottom=Side(style="thin", color=BORDER))

    # ── Bloc « Résumé du portefeuille » ────────────────────────────────────
    row = 10
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = f"  {_t(lang, 'summary').upper()}"
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24

    # Les exports Excel tolèrent l'absence de taux : on affiche la valeur brute
    # plutôt que de faire planter tout le rapport (strict=False).
    initial = convert_amount(db, float(investment.initial_capital or 0), inv_ccy, display_ccy, strict=False) if db else float(investment.initial_capital or 0)
    current = convert_amount(db, float(investment.current_value or 0), inv_ccy, display_ccy, strict=False) if db else float(investment.current_value or 0)

    # Bénéfice réalisé = Σ gains − Σ (pertes + frais), apports exclus.
    pnl = 0.0
    net_deposits = 0.0
    for tx in transactions:
        tx_ccy = (getattr(tx, "currency", None) or "HTG").upper()
        amt = convert_amount(db, float(tx.amount or 0), tx_ccy, display_ccy, strict=False) if db else float(tx.amount or 0)
        if tx.type == "gain":
            pnl += amt
        elif tx.type in ("loss", "fee"):
            pnl -= amt
        elif tx.type == "deposit":
            net_deposits += amt
        elif tx.type == "withdrawal":
            net_deposits -= amt
    roi = (pnl / initial * 100) if initial else 0

    # 3 colonnes pour un rendu plus moderne : label (A-B) · valeur (C-D) · label (E) · valeur (F)
    summary_rows = [
        (_t(lang, "initial"),     f"{initial:,.2f} {display_ccy}",  _t(lang, "entry_date"), str(investor.entry_date)),
        (_t(lang, "current"),     f"{current:,.2f} {display_ccy}",  _t(lang, "status"),     (investor.status or "").upper()),
        (_t(lang, "pnl"),         f"{pnl:+,.2f} {display_ccy}",     "",                     ""),
        (_t(lang, "roi"),         f"{roi:+.2f}%",                   "",                     ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(summary_rows):
        r = row + 1 + i
        bg = LIGHTER if i % 2 == 0 else WHITE
        _paint_row(ws, r, 1, 6, bg)
        # gauche
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        a = ws.cell(row=r, column=1, value=f"  {l1}")
        a.font = Font(bold=True, color=TEXT_MUTED, size=10)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        b = ws.cell(row=r, column=3, value=v1)
        b.font = Font(bold=True, color=TEXT_DARK, size=11)
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # droite
        c = ws.cell(row=r, column=5, value=l2)
        c.font = Font(bold=True, color=TEXT_MUTED, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        d = ws.cell(row=r, column=6, value=v2)
        d.font = Font(bold=True, color=TEXT_DARK, size=11)
        d.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
    # bordure sous le bloc résumé
    for col in range(1, 7):
        ws.cell(row=row + len(summary_rows), column=col).border = Border(bottom=Side(style="thin", color=BORDER))

    # ── Historique des transactions ────────────────────────────────────────
    row = row + len(summary_rows) + 2  # espace
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = f"  {_t(lang, 'history').upper()}"
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24

    header_row = row + 1
    headers = [
        _t(lang, "date"),
        _t(lang, "type"),
        f"{_t(lang, 'amount')} ({display_ccy})",
        _t(lang, "original"),
        _t(lang, "orig_ccy"),
        _t(lang, "description"),
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=PRIMARY_SOFT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border(PRIMARY_SOFT)
    ws.row_dimensions[header_row].height = 22

    # Ligne synthétique « INITIAL » : le capital initial est traité comme une
    # première transaction d'ouverture. Ça répond à la remarque utilisateur :
    # « je ne vois même pas le montant comme apport initial comme une transaction ».
    initial_original = float(investment.initial_capital or 0)
    initial_converted = initial
    synthetic_initial = {
        "date": str(getattr(investment, "start_date", None) or investor.entry_date),
        "type": "initial",
        "converted": initial_converted,
        "original": initial_original,
        "orig_ccy": inv_ccy,
        "description": _t(lang, "initial_deposit_desc"),
    }

    tx_rows = [synthetic_initial] + [
        {
            "date": str(tx.transaction_date),
            "type": (tx.type or "").lower(),
            "converted": (convert_amount(db, float(tx.amount or 0), (getattr(tx, "currency", None) or "HTG").upper(), display_ccy, strict=False) if db else float(tx.amount or 0)),
            "original": float(tx.amount or 0),
            "orig_ccy": (getattr(tx, "currency", None) or "HTG").upper(),
            "description": tx.description or _t(lang, "na"),
        }
        for tx in transactions
    ]

    for i, tx in enumerate(tx_rows):
        r = header_row + 1 + i
        style = TX_TYPE_STYLES.get(tx["type"], {"bg": WHITE if i % 2 else LIGHTER, "fg": TEXT_DARK})
        bg = style["bg"]
        fg = style["fg"]
        # Painted background on full row for strong visual grouping
        _paint_row(ws, r, 1, 6, bg)

        # Date
        a = ws.cell(row=r, column=1, value=tx["date"])
        a.font = Font(color=TEXT_DARK, size=10)
        a.alignment = Alignment(horizontal="left", indent=1, vertical="center")

        # Type — badge-like (coloured bold label)
        b = ws.cell(row=r, column=2, value=_type_label(tx["type"], lang))
        b.font = Font(bold=True, color=fg, size=10)
        b.alignment = Alignment(horizontal="left", indent=1, vertical="center")

        # Converted amount (signed for clarity)
        signed = tx["converted"]
        if tx["type"] in ("loss", "fee", "withdrawal"):
            signed = -abs(signed)
        elif tx["type"] in ("gain",):
            signed = abs(signed)
        c = ws.cell(row=r, column=3, value=signed)
        c.font = Font(bold=True, color=fg, size=10)
        c.number_format = '#,##0.00;[Red]-#,##0.00'
        c.alignment = Alignment(horizontal="right", vertical="center")

        # Original amount + currency
        d = ws.cell(row=r, column=4, value=tx["original"])
        d.font = Font(color=TEXT_DARK, size=10)
        d.number_format = '#,##0.00'
        d.alignment = Alignment(horizontal="right", vertical="center")

        e = ws.cell(row=r, column=5, value=tx["orig_ccy"])
        e.font = Font(color=TEXT_MUTED, size=10)
        e.alignment = Alignment(horizontal="center", vertical="center")

        f = ws.cell(row=r, column=6, value=tx["description"])
        f.font = Font(color=TEXT_DARK, size=10)
        f.alignment = Alignment(horizontal="left", indent=1, vertical="center", wrap_text=True)

        # Thin bottom border between rows
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = Border(bottom=Side(style="thin", color=BORDER))
        ws.row_dimensions[r].height = 20

    # ── Légende des couleurs ───────────────────────────────────────────────
    last_tx_row = header_row + len(tx_rows)
    legend_row = last_tx_row + 2
    ws.merge_cells(f"A{legend_row}:F{legend_row}")
    c = ws[f"A{legend_row}"]
    c.value = f"  {_t(lang, 'legend').upper()}"
    c.font = Font(bold=True, color=TEXT_MUTED, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[legend_row].height = 16

    legend_types = ["initial", "deposit", "withdrawal", "gain", "loss", "fee"]
    for idx, t in enumerate(legend_types):
        meta = TX_TYPE_STYLES[t]
        col = idx + 1
        cell = ws.cell(row=legend_row + 1, column=col, value=f"  ● {_type_label(t, lang)}")
        cell.fill = PatternFill(fill_type="solid", fgColor=meta["bg"])
        cell.font = Font(bold=True, color=meta["fg"], size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _thin_border(BORDER)
    ws.row_dimensions[legend_row + 1].height = 22

    # ── Performances (optionnel) ───────────────────────────────────────────
    if performances:
        perf_start = legend_row + 3
        ws.merge_cells(f"A{perf_start}:F{perf_start}")
        c = ws[f"A{perf_start}"]
        c.value = f"  {_t(lang, 'performances').upper()}"
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[perf_start].height = 24

        ph = perf_start + 1
        perf_headers = [_t(lang, "period"), _t(lang, "start"), _t(lang, "end"), _t(lang, "roi_pct"), _t(lang, "gross_gain"), _t(lang, "drawdown")]
        for col, h in enumerate(perf_headers, start=1):
            c = ws.cell(row=ph, column=col, value=h)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill(fill_type="solid", fgColor=PRIMARY_SOFT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border(PRIMARY_SOFT)
        ws.row_dimensions[ph].height = 22

        for i, p in enumerate(performances):
            pr = ph + 1 + i
            bg = LIGHTER if i % 2 == 0 else WHITE
            _paint_row(ws, pr, 1, 6, bg)
            vals = [
                p.period_type,
                str(p.period_start) if p.period_start else _t(lang, "na"),
                str(p.period_end) if p.period_end else _t(lang, "na"),
                float(p.roi_pct) if getattr(p, "roi_pct", None) else 0.0,
                float(p.gross_gain) if getattr(p, "gross_gain", None) else 0.0,
                float(p.max_drawdown_pct) if getattr(p, "max_drawdown_pct", None) else 0.0,
            ]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=pr, column=col, value=val)
                c.font = Font(color=TEXT_DARK, size=10)
                c.border = Border(bottom=Side(style="thin", color=BORDER))
                if col in (4, 5, 6):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            ws.row_dimensions[pr].height = 20

    # ── Pied de page ───────────────────────────────────────────────────────
    footer_row = (ws.max_row or 1) + 2
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = f"© {datetime.now().year} {company_name} — Document généré automatiquement"
    c.font = Font(color=TEXT_MUTED, italic=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
